#! /usr/bin/python3
# invertSpreadsheet.py -> inverts the rows and columns of a spreadsheet

import openpyxl
import os
import sys

# Print usage info if multiple arguments not provided
if len(sys.argv) < 2:
    print(r'''Usage:
        python3 invertSpreadsheet.py [file]
        python3 invertSpreadsheet.py [file] [output]
            ''')
    exit()

# Get workbook
print('Opening file...')
wb = sys.argv[1]
wb = os.path.abspath(wb)
originalWorkbook = openpyxl.load_workbook(wb)
sheet = originalWorkbook.active

# Read in original data as a list of lists
print('Reading in data')
sheetData = []
max_row = sheet.max_row
max_column = sheet.max_column
for row in range(1, max_row + 1):
    rowData = []
    for column in range(1, max_column + 1):
        cell_value = sheet.cell(row=row, column=column).value
        rowData.append(cell_value)
    sheetData.append(rowData)

# Create new workbook
print('Creating output sheet')
newWorkbook = openpyxl.Workbook()
outputSheet = newWorkbook.active

''' DOESN'T RUN - INDEX OUT OF RANGE'''
# Populate output sheet with inverted version of original data
print('Transposing data to output sheet')
for x in range(1, sheet.max_row + 1):
    for y in range(1, sheet.max_column + 1):
        outputSheet.cell(row=y, column=x).value = sheetData[y][x]
        
# Save the file
if len(sys.argv) < 3:
    output.save('inverted' + os.path.basename(wb))
    print('File saved as inverted' + os.path.basename(wb))
else:
    output.save(sys.argv[2])
    print('File saved as ' + sys.argv[2])

print('Done')
