#! python3
# ABS_12_updatepProduce.py -> Corrects costs in produce spreadsheet

import openpyxl

print('Opening workbook...')
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

print('Reading rows...')
    
# Loop through the rows and updated the prices
for rowNum in range(2, sheet.max_row + 1):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

print('Saving updated version...')
wb.save('updatedProduceSales.xlsx')
print('Done')
