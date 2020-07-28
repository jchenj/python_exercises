#! /usr/bin/python3
# ABS_14_excelToCsv.py -> reads all Excel files in current working directory
# and outputs them as CSV files

import csv
import os
import openpyxl

for excelFile in os.listdir('.'):
    # Skip non-xlsx files
    if not excelFile.endswith('.xlsx'):
        continue
    print('Opening file ' + excelFile + '...')
    
    # Load xlsx file
    wb = openpyxl.load_workbook(excelFile)     
    for sheetName in wb.sheetnames:
        # Loop through every sheet in the workbook
        sheet = wb[sheetName]
        # Create the CSV filename from the Excel filename and sheet title
        excelFilename = excelFile.rstrip('.xlsx')
        sheetTitle = sheet.title
        csvFile = open(excelFilename + '_' + sheetTitle + '.csv', 'w', newline = '')
        # Create the csv.writer object for this CSV file
        csvWriter = csv.writer(csvFile) 
        # Loop through every row in the sheet
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []   # Append each cell to this list
            # Loop through each cell in the row
            for colNum in range(1, sheet.max_column + 1):
                cellValue = sheet.cell(row=rowNum, column=colNum).value
                # Append each cell's data to rowData
                rowData.append(cellValue)
            # Write the rowData list to the CSV file
            csvWriter.writerow(rowData)
    csvFile.close()


        
